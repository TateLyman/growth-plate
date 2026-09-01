"""Find TET1/TET2/TET3 (and controls) in the kosmicki 1.45M-exome height tables.

The question: auranofin's 2-OG pocket is 9/9 identical across TET1/2/3, so a
pocket-directed agent is effectively PAN-TET. What does human genetics predict
for pan-TET partial loss, rather than for TET1 alone?
"""
import openpyxl

P = '/home/user/growth-plate/atlas/data/supplied_2026_08_12/kosmicki2026_supp_tables_S1_S29.xlsx'
wb = openpyxl.load_workbook(P, read_only=True, data_only=True)

TARGETS = {'TET1', 'TET2', 'TET3', 'ACAN', 'FGFR3', 'NPR2', 'IHH'}

for sn in wb.sheetnames:
    ws = wb[sn]
    rows = ws.iter_rows(values_only=True)
    header = None
    hits = []
    for i, r in enumerate(rows):
        if r is None:
            continue
        cells = [str(c) if c is not None else '' for c in r]
        if header is None:
            joined = ' '.join(cells).lower()
            if any(k in joined for k in ('gene', 'symbol')) and i < 6:
                header = cells
                continue
        if header is None:
            continue
        for c in cells[:4]:
            if c.strip() in TARGETS:
                hits.append((c.strip(), cells))
                break
        if i > 60000:
            break
    if hits:
        print('\n===== %s  (%d target rows) =====' % (sn, len(hits)))
        print('HEADER:', [h for h in header][:20])
        for g, cells in hits[:40]:
            print('  ', cells[:20])
