import openpyxl, sys
wb = openpyxl.load_workbook(sys.argv[1], read_only=True, data_only=True)
print("sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print("\n===== %s  (%s x %s) =====" % (name, ws.max_row, ws.max_column))
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 4: break
        print("   ", [str(c)[:34] if c is not None else '' for c in row[:12]])
