"""GSE114919, done properly: per-sample values, Welch t-tests, both species.

The confound-free contrast: SLOW-growing plate (phalanx/finger) vs FAST-growing
plate (tibia) at MATCHED age and MATCHED zone. Values are log2, so the statistic
is a DIFFERENCE (delta log2), not a ratio.
"""
import openpyxl, re, statistics as st, math

GENES = ['Tet1', 'Tet2', 'Tet3', 'Ihh', 'Col10a1', 'Acan', 'Col2a1',
         'Mki67', 'Pthlh', 'Sox9', 'Fgfr3', 'Igf1', 'Nppc', 'Npr2']

RATMAP = {}
_wb = openpyxl.load_workbook('/home/user/gp_data/gse114919_ratnames.xlsx', read_only=True, data_only=True)
for _r in _wb[_wb.sheetnames[0]].iter_rows(values_only=True):
    if _r and _r[0] is not None and str(_r[0]).strip().isdigit():
        RATMAP[int(str(_r[0]).strip())] = str(_r[1]).strip()
print('rat sample labels:', sorted(set(re.sub(r'\d+$', '', v) for v in RATMAP.values())))


def welch(a, b):
    if len(a) < 2 or len(b) < 2:
        return float('nan'), float('nan')
    ma, mb = st.mean(a), st.mean(b)
    va, vb = st.variance(a), st.variance(b)
    se = math.sqrt(va / len(a) + vb / len(b))
    if se == 0:
        return float('nan'), float('nan')
    t = (ma - mb) / se
    df = (va / len(a) + vb / len(b)) ** 2 / (
        (va / len(a)) ** 2 / (len(a) - 1) + (vb / len(b)) ** 2 / (len(b) - 1))
    # two-sided p via incomplete beta
    x = df / (df + t * t)

    def betacf(a_, b_, x_):
        MAXIT, EPS, FPMIN = 200, 3e-12, 1e-300
        qab, qap, qam = a_ + b_, a_ + 1.0, a_ - 1.0
        c, d = 1.0, 1.0 - qab * x_ / qap
        if abs(d) < FPMIN: d = FPMIN
        d = 1.0 / d; h = d
        for m in range(1, MAXIT + 1):
            m2 = 2 * m
            aa = m * (b_ - m) * x_ / ((qam + m2) * (a_ + m2))
            d = 1.0 + aa * d
            if abs(d) < FPMIN: d = FPMIN
            c = 1.0 + aa / c
            if abs(c) < FPMIN: c = FPMIN
            d = 1.0 / d; h *= d * c
            aa = -(a_ + m) * (qab + m) * x_ / ((a_ + m2) * (qap + m2))
            d = 1.0 + aa * d
            if abs(d) < FPMIN: d = FPMIN
            c = 1.0 + aa / c
            if abs(c) < FPMIN: c = FPMIN
            d = 1.0 / d; de = d * c; h *= de
            if abs(de - 1.0) < EPS: break
        return h

    def betai(a_, b_, x_):
        if x_ <= 0: return 0.0
        if x_ >= 1: return 1.0
        lb = (math.lgamma(a_ + b_) - math.lgamma(a_) - math.lgamma(b_)
              + a_ * math.log(x_) + b_ * math.log(1.0 - x_))
        if x_ < (a_ + 1.0) / (a_ + b_ + 2.0):
            return math.exp(lb) * betacf(a_, b_, x_) / a_
        return 1.0 - math.exp(lb) * betacf(b_, a_, 1.0 - x_) / b_
    return t, betai(df / 2.0, 0.5, x)


def load(path, ratstyle):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h) if h is not None else '' for h in next(it)]
    # find which header cells are real sample names
    idx = {}
    for i, c in enumerate(hdr):
        s = c.strip()
        k = None
        m = re.search(r'(\d+)w(Ph|T)_(PZ|HZ)', s, re.I)
        if m:
            k = (m.group(2).upper(), int(m.group(1)), m.group(3).upper())
        if ratstyle:
            m = re.search(r'JL-\d+-(\d+)_S', s)
            if m:
                lab = RATMAP.get(int(m.group(1)), '')
                mm = re.search(r'(T|Ph)(\d+)wk\s*(PZ|HZ)', lab, re.I)
                if mm:
                    k = (mm.group(1).upper(), int(mm.group(2)), mm.group(3).upper())
        if k:
            idx.setdefault(k, []).append(i)
    rows = {}
    LOWER = {g.lower(): g for g in GENES}
    symcols = (1, 0) if ratstyle else (0, 1)
    for r in it:
        if not r:
            continue
        for sc in symcols:
            if sc < len(r) and r[sc] is not None:
                g = LOWER.get(str(r[sc]).strip().lower())
                if g:
                    rows.setdefault(g, []).append(r)
                    break
    return idx, rows


def report(path, ratstyle, label):
    print('\n' + '=' * 84)
    print(label)
    idx, rows = load(path, ratstyle)
    if not idx:
        print('  no groups'); return
    print('  groups:', {'%s%dwk_%s' % k: len(v) for k, v in sorted(idx.items())})
    contrasts = sorted({(a, z) for (b, a, z) in idx if ('T', a, z) in idx and ('PH', a, z) in idx})
    print('  confound-free SLOW-vs-FAST contrasts available:', contrasts)
    for (age, zone) in contrasts:
        print('\n  --- %dwk %s :  phalanx(SLOW) minus tibia(FAST), log2 ---' % (age, zone))
        print('  %-9s %9s %9s %9s %8s %9s' % ('gene', 'SLOW', 'FAST', 'dlog2', 'x-fold', 'p'))
        for g in GENES:
            if g not in rows:
                continue
            rr = rows[g][0]

            def vals(k):
                out = []
                for i in idx[k]:
                    try:
                        out.append(float(rr[i]))
                    except (TypeError, ValueError, IndexError):
                        pass
                return out
            s, f = vals(('PH', age, zone)), vals(('T', age, zone))
            if len(s) < 2 or len(f) < 2:
                continue
            d = st.mean(s) - st.mean(f)
            t, p = welch(s, f)
            star = '  <<<' if (p == p and p < 0.05) else ''
            print('  %-9s %9.2f %9.2f %+9.2f %8.2f %9.4f%s'
                  % (g, st.mean(s), st.mean(f), d, 2 ** d, p, star))
        print('  (positive dlog2 = HIGHER in the SLOW-growing plate)')


report('/home/user/gp_data/gse114919_mouse.xlsx', False, 'MOUSE - finger(slow) vs tibia(fast)')
report('/home/user/gp_data/gse114919_rat.xlsx', True, 'RAT - phalanx(slow) vs tibia(fast)')
