"""Quantify R159's ACAN CONFLICT from Smeriglio 2020's own supplementary DE tables.

R159: "the precise mechanism by which TET1 loss could hurt height is the
downregulation of the one gene whose loss costs more height than any other"
(ACAN pLoF = -14.1 cm/allele). It was called "the sharpest new hole" and never
quantified. These are the authors' own tables: shRNA KD (ATDC5) and Tet1 KO.
"""
import openpyxl, re, statistics as st

F = '/home/user/gp_data/smersupp/JBM4-4-e10383-s005.xlsx'
wb = openpyxl.load_workbook(F, read_only=True, data_only=True)

TARGETS = {'acan', 'col2a1', 'col10a1', 'sox9', 'col9a1', 'col9a2', 'col11a1',
           'ihh', 'pthlh', 'tet1', 'tet2', 'tet3', 'fgfr3', 'npr2', 'nppc',
           'mki67', 'sox5', 'sox6', 'runx2', 'mmp13', 'spp1'}

# ---------- KO sheet ----------
ws = wb['KO differentially expressed']
rows = list(ws.iter_rows(values_only=True))
hdr = [str(x) if x else '' for x in rows[0]]
print('KO sheet header:', hdr)
ko = []
for r in rows[1:]:
    if not r or r[0] is None:
        continue
    ko.append(r)
print('KO differentially expressed genes listed:', len(ko))

fcs = []
for r in ko:
    try:
        fcs.append(float(r[3]))
    except (TypeError, ValueError, IndexError):
        pass
print('  fold-change range: %.3f to %.3f' % (min(fcs), max(fcs)))
print('  n with FC<1 (DOWN in KO): %d    n with FC>1 (UP in KO): %d'
      % (sum(1 for x in fcs if x < 1), sum(1 for x in fcs if x > 1)))
sox9 = sum(1 for r in ko if len(r) > 4 and str(r[4]).strip().lower() == 'yes')
hmc = sum(1 for r in ko if len(r) > 5 and str(r[5]).strip().lower() == 'yes')
print('  annotated SOX9 target: %d   annotated 5hmC Day20: %d' % (sox9, hmc))

print('\n  --- TARGET GENES IN THE KO DE LIST ---')
found = False
for r in ko:
    g = str(r[0]).strip()
    if g.lower() in TARGETS:
        found = True
        print('   %-10s FC %8.3f   p %-12s q %-12s SOX9=%s 5hmC=%s'
              % (g, float(r[3]), r[1], r[2], r[4] or '-', r[5] or '-'))
if not found:
    print('   *** NONE OF Acan, Col2a1, Col10a1, Sox9, Ihh, Pthlh, Fgfr3, Npr2 APPEARS')
    print('   *** IN THE KO DIFFERENTIALLY-EXPRESSED LIST AT ALL. ***')

# ---------- KD sheet ----------
ws = wb['KD differentially expressed']
rows = list(ws.iter_rows(values_only=True))
hdr = [str(x) if x else '' for x in rows[0]]
print('\nKD sheet header:', hdr[:14])
kd = [r for r in rows[1:] if r and r[2] is not None]
print('KD differentially expressed probes listed:', len(kd))
print('\n  --- TARGET GENES IN THE KD DE LIST ---')
found = False
for r in kd:
    g = str(r[2]).strip()
    if g.lower() in TARGETS:
        found = True
        vals = [x for x in r[6:14]]
        print('   %-10s %s' % (g, ' '.join('%s' % (('%.1f' % v) if isinstance(v, (int, float)) else str(v)[:9]) for v in vals)))
if not found:
    print('   *** NONE OF THE CARTILAGE PANEL APPEARS IN THE KD DE LIST EITHER. ***')

# how big is the KD effect overall?
print('\n  --- other supplementary sheets ---')
import glob
for f in sorted(glob.glob('/home/user/gp_data/smersupp/*.xlsx')):
    w = openpyxl.load_workbook(f, read_only=True, data_only=True)
    print('   %-18s %s' % (f.split('/')[-1], w.sheetnames))
