#!/usr/bin/env python3
"""Compute the YIELD - bone elongation obtained per resting-zone chondrocyte spent -
from the raw per-animal values in lui2018's S1 Data (PLoS Biology, CC-BY).

WHY THIS EXISTS
---------------
`g_l2_raise_the_yield_per_progenitor` records that the yield has never been measured
in any species, twenty years after schrier2006 named it and twelve after nilsson2014
restated it. Both of those papers report their terms only in raster figures. lui2018
publishes the underlying per-animal values as an open-access supplementary workbook,
and it happens to contain both terms in the same mice at the same ages:

  numerator   calcein-labelled bone growth rate, um/day        (S1 Data sheet Fig1C)
  denominator resting zone cell count per 500 um plate width   (S1 Data sheet Fig2B-G)

THE CONSTRUCTION
----------------
A yield is a flux over a flux, so a single-timepoint ratio of two standing stocks is
the WRONG quantity and would produce a confident wrong number. Over an interval:

    yield = (um of bone elongated between t1 and t2)
            / (resting zone cells lost between t1 and t2, per 500 um width)

Physically: take a 500 um-wide slab of growth plate. It elongates by L um over the
interval while its resting zone loses N cells. L/N is um of bone per resting zone cell.

WHAT THIS IS NOT
----------------
This is a RE-ANALYSIS of another group's published values, not a measurement. It is
graded as such wherever it enters the atlas. Its specific limits are printed with the
results and repeated in the node - the most important being that a NET change in a
standing stock is not gross consumption, so every number here is an UPPER BOUND on the
true yield.

Data: doi:10.1371/journal.pbio.2005263 S1 Data, CC-BY 4.0.
"""
import sys
import numpy as np
import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else "lui2018_S1Data.xlsx"
BONES = ["femur", "tibia", "metacarpal", "phalanx"]
RNG = np.random.default_rng(20260807)
NBOOT = 20000


def age_weeks(a):
    """Map the workbook's age labels to weeks. E17/NB are pre-resting-zone."""
    if a is None:
        return None
    s = str(a).strip().lower()
    if s in ("e17", "e17.5", "nb"):
        return None
    if s.endswith("wk"):
        s = s[:-2]
    try:
        return float(s)
    except ValueError:
        return None


def _block(rows, label):
    """Resolve a labelled block to (age_col, {bone: col}).

    The workbook is not positionally consistent - in sheet Fig2B-G the block
    label sits directly over its first bone column, in Fig1C it sits one column
    to the right of the block's own Age column. So locate the label in the top
    header row, then find the first 'femur' at or after it in the second header
    row, and take the nearest preceding 'Age' as that block's age column.
    """
    top, sub = rows[0], rows[1]
    li = top.index(label)
    fem = next(i for i in range(li - 1, len(sub))
               if str(sub[i]).strip().lower() == "femur")
    age = max(i for i in range(fem) if str(sub[i]).strip().lower() == "age")
    cols = {}
    for b in BONES:
        cols[b] = next(i for i in range(fem, len(sub))
                       if str(sub[i]).strip().lower() == b)
    return age, cols


def _harvest(rows, label):
    age_col, cols = _block(rows, label)
    out = {}
    for r in rows[2:]:
        if age_col >= len(r):
            continue
        w = age_weeks(r[age_col])
        if w is None:
            continue
        for bone, c in cols.items():
            v = r[c] if c < len(r) else None
            if isinstance(v, (int, float)) and v > 0:
                out.setdefault((bone, w), []).append(float(v))
    return out


def load(xlsx):
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    # resting zone cell count, per 500 um plate width, mouse
    rz = _harvest([list(r) for r in wb["Fig2B-G"].iter_rows(values_only=True)],
                  "Resting Zone Cell Count")
    # calcein-labelled bone growth rate, um/day, mouse
    gr = _harvest([list(r) for r in wb["Fig1C"].iter_rows(values_only=True)],
                  "Calcein labeled bone growth")
    return rz, gr


def yield_point(rz1, rz2, g1, g2, days):
    """um elongated over the interval / RZ cells lost over the interval."""
    grown = 0.5 * (np.mean(g1) + np.mean(g2)) * days      # trapezoid on the rate
    lost = np.mean(rz1) - np.mean(rz2)                    # NET loss, per 500 um
    if lost <= 0:
        return grown, lost, None
    return grown, lost, grown / lost


def boot(rz1, rz2, g1, g2, days):
    out = []
    for _ in range(NBOOT):
        b = lambda a: RNG.choice(a, size=len(a), replace=True)
        _, _, y = yield_point(b(rz1), b(rz2), b(g1), b(g2), days)
        if y is not None:
            out.append(y)
    if len(out) < NBOOT * 0.5:
        return None, None, len(out) / NBOOT
    return (float(np.percentile(out, 2.5)), float(np.percentile(out, 97.5)),
            len(out) / NBOOT)


def main():
    rz, gr = load(XLSX)
    ages = sorted({w for (_, w) in gr})
    print("=" * 78)
    print("YIELD: um of bone elongation per resting-zone chondrocyte spent")
    print("re-analysis of lui2018 S1 Data (PLoS Biology, CC-BY); mouse")
    print("=" * 78)
    print(f"calcein ages available: {ages}")
    print(f"bootstrap resamples: {NBOOT}\n")

    hdr = (f"{'bone':<11}{'interval':<11}{'grown':>9}{'RZ lost':>9}"
           f"{'YIELD':>9}{'95% CI':>21}{'n':>9}")
    print(hdr)
    print("-" * len(hdr))

    for bone in BONES:
        for t1, t2 in zip(ages, ages[1:]):
            k1, k2 = (bone, t1), (bone, t2)
            if not all(k in rz for k in (k1, k2)):
                continue
            if not all(k in gr for k in (k1, k2)):
                continue
            if t1 == 1.0:
                # EXCLUDED, not merely flagged. lui2018 defines the resting zone's
                # upper margin differently at 1 week (the FUTURE secondary
                # ossification centre) than from 2 weeks on (the LOWER MARGIN of the
                # actual SOC). The 1-week count therefore includes tissue the SOC
                # subsequently occupies, and the apparent 50-60% fall to 2 weeks is
                # substantially the SOC forming rather than progenitors being spent.
                # An inflated denominator makes the yield look small; the four
                # 1-2wk values all landed at 20-28 for exactly that reason.
                print(f"{bone:<11}{f'{t1:g}-{t2:g}wk':<11}"
                      f"{'EXCLUDED - resting zone redefined at 2wk (SOC)':>58}")
                continue
            days = (t2 - t1) * 7.0
            grown, lost, y = yield_point(rz[k1], rz[k2], gr[k1], gr[k2], days)
            ns = f"{len(rz[k1])}/{len(rz[k2])}"
            if y is None:
                print(f"{bone:<11}{f'{t1:g}-{t2:g}wk':<11}{grown:>9.0f}"
                      f"{lost:>9.1f}{'n/d':>9}{'RZ did not fall':>21}{ns:>9}")
                continue
            lo, hi, frac = boot(rz[k1], rz[k2], gr[k1], gr[k2], days)
            ci = f"[{lo:.0f}, {hi:.0f}]" if lo is not None else "unstable"
            print(f"{bone:<11}{f'{t1:g}-{t2:g}wk':<11}{grown:>9.0f}"
                  f"{lost:>9.1f}{y:>9.0f}{ci:>21}{ns:>9}")

    print("\nunits: grown = um elongated over the interval (trapezoid on calcein rate)")
    print("       RZ lost = NET fall in resting zone cells per 500 um plate width")
    print("       YIELD = um of bone per resting zone cell lost, per 500 um width")
    print("""
LIMITS - every one of these is carried into the atlas with the number:
  1 NET, NOT GROSS. A fall in a standing stock is (cells leaving) minus (cells
    self-renewing). Gross consumption is therefore >= the denominator used here,
    so EVERY VALUE IS AN UPPER BOUND on the true yield.
  2 DENSITY, NOT COUNT. Cells per 500 um width is a density. The plate widens with
    age, so total resting zone cells could rise while this falls.
  3 THE 1-2wk INTERVAL IS EXCLUDED, not down-weighted. lui2018 defines the
    resting zone's upper margin as the FUTURE secondary ossification centre at
    1 week and as the LOWER MARGIN OF THE ACTUAL SOC from 2 weeks on, so the
    1-week count includes tissue the SOC later occupies. The 50-60% fall from
    1 to 2 weeks is substantially the SOC forming, not progenitors being spent.
  4 RATE INTERPOLATION. Calcein gives an instantaneous rate at each age; the
    trapezoid assumes it changes linearly in between.
  5 MOUSE, and mouse femur and tibia do not fuse. A yield curve in a bone that
    never runs out is not a spend-to-exhaustion curve. The metacarpal and phalanx,
    which DO fuse, are the informative pair.
  6 TWO SEPARATE COHORTS. Calcein and histology animals are not the same mice, so
    the numerator and denominator are matched by age and bone, not by individual.
""")


if __name__ == "__main__":
    main()
