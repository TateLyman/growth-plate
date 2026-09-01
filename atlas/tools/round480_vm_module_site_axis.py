#!/usr/bin/env python3
"""
R480 - DOES THE MATRIX MODULE TRACK THE MATRIX FRACTION? THE ONE SITE-x-ZONE DATASET,
BOTH SPECIES, WITH R312's FOUR MEASURED CONFOUNDS AS CONTROLS.

THE PREDICTION, WRITTEN BEFORE LOOKING. The identity is dL/dt = flux x v(d) with
v(d) = v(c) + v(m). Wilsman 1996 and Breur 1997 agree that as a plate slows, the
MATRIX FRACTION of the terminal domain RISES - 31.5 per cent in the fastest rat plate
to 49.7 per cent in the slowest - while terminal cell volume falls 3.6 to 3.9-fold and
matrix per cell falls only 1.7 to 1.9-fold. GSE114919's site contrast is 1-week TIBIA
(fast) against 1-week PHALANX (slow) in the same animals at the same age, zone-resolved,
in mouse AND rat.

  IF the matrix fraction rises as the plate slows, the matrix-synthesis module should sit
  HIGHER in the PHALANX than the tibia RELATIVE TO THE GENOME-WIDE BACKGROUND, and the
  v(c) module (F-R185's 23 h_term genes) should not, or should run the other way.

WHY THIS IS NOT A DISCOVERY SCREEN. R312 measured four confounds in exactly this dataset
and CLAUDE.md's instruction is explicit: "Use it to adjudicate genes you already have
reasons for, never as a discovery screen." Muscle contamination alone scores 6.15 in the
1-week tibia against 1.82 in the phalanx and reproduces the capacity pattern by itself.
So every confound panel is scored alongside, and every module is reported as a PERCENTILE
of the genome-wide distribution rather than as a raw direction - CORR-329's base-rate rule.

WHAT A BULK CONTRAST CANNOT DO, STATED FIRST. Transcript abundance per unit RNA is not
matrix volume per cell. If the phalanx terminal cell is smaller, per-cell transcript for
everything differs. This test can show that the matrix module is DIFFERENTIALLY INVESTED
between a fast and a slow plate; it cannot measure v(m). Nothing here is a volume.

Output: atlas/data/round480/vm_module_site_axis.json
"""
import json
import os
import statistics
import sys

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                "..", "quant", "notebooks", "p10_02_gse114919_age_zone_site"))
from analysis import load, groups, contrast  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE = os.path.join(ROOT, "data", "supplied_2026_08_13")
OUT = os.path.join(ROOT, "data", "round480")

MODULES = {
    # ---- v(m): the machinery that makes the matrix half of the domain
    "vm_core_protein": ["ACAN", "VCAN", "HAPLN1", "BGN", "DCN", "FMOD", "LUM",
                        "PRELP", "EPYC", "OGN", "ASPN", "CHAD", "COMP", "MATN1", "MATN3"],
    "vm_collagen": ["COL2A1", "COL9A1", "COL9A2", "COL9A3", "COL11A1", "COL11A2",
                    "COL10A1", "COL27A1"],
    "vm_gag_synthesis": ["XYLT1", "XYLT2", "B4GALT7", "B3GALT6", "B3GAT3", "FAM20B",
                         "CSGALNACT1", "CSGALNACT2", "CHSY1", "CHSY3", "CHPF", "CHPF2"],
    "vm_sulfation": ["CHST3", "CHST11", "CHST12", "CHST13", "CHST14", "CHST7", "UST",
                     "PAPSS1", "PAPSS2", "SLC35B2", "SLC35B3", "SLC26A2", "IMPAD1", "BPNT1"],
    "vm_collagen_folding": ["SERPINH1", "P4HA1", "P4HA2", "P4HB", "P3H1", "P3H2", "P3H3",
                            "CRTAP", "PPIB", "PLOD1", "PLOD2", "PLOD3", "FKBP10", "FKBP14"],
    "vm_secretory_export": ["MIA3", "MIA2", "SEC23A", "SEC23B", "SEC24D", "SEC24C",
                            "SEC13", "SEC31A", "SAR1A", "SAR1B", "CREB3L2", "CREB3L1",
                            "TRAPPC2", "KDELR1", "KDELR2", "TMED9", "TMED2", "TMED10", "SURF4"],
    "vm_golgi_sugar": ["SLC35D1", "SLC35A1", "SLC35A2", "SLC35A3", "UGDH", "UXS1",
                       "UGP2", "GALE", "CANT1", "ENTPD5", "TMEM165"],
    # ---- v(c): F-R185's 23-gene h_term screen, which is the comparator
    "vc_fr185_screen": ["PAPPA", "PAPPA2", "STC2", "IGFBP4", "IGFBP5", "IGF1", "IGF1R",
                        "NPPC", "NPR2", "NPR3", "OSTN", "MME", "IHH", "SOX9", "RUNX2",
                        "GLI2", "PTH1R", "HDAC4", "SMO", "MTOR", "FGFR3", "INPPL1"],
    # ---- R312's four MEASURED confounds in this exact dataset
    "confound_muscle": ["MYH8", "MYH4", "MYH3", "ACTA1", "MYL1", "TNNT3", "TNNC2",
                        "DES", "CKM", "MYOG"],
    "confound_distal_hox": ["HOXA13", "HOXD13", "HOXA11", "HOXD11", "HOXD12", "HOXA10"],
    "confound_bone": ["BGLAP", "SOST", "MEPE", "IBSP", "DMP1", "SP7", "COL1A1"],
    "confound_marrow_immune": ["MS4A1", "CD74", "CCL6", "SIGLEC1", "PTPRC", "HBB",
                               "LYZ2", "CTSS"],
    # ---- housekeeping negative control: must sit at the median
    "control_housekeeping": ["GAPDH", "ACTB", "PPIA", "TUBB5", "RPL13A", "B2M",
                             "TBP", "HPRT1", "PGK1", "SDHA"],
}


def lookup(data, sym):
    """Case-insensitive symbol match; mouse/rat matrices use Title case."""
    for cand in (sym, sym.capitalize(), sym.title(), sym.lower(), sym.upper()):
        if cand in data:
            return cand
    return None


def score(con, data, gl):
    """Return per-gene log2 site differences for a module, and its summary."""
    hits, vals = [], []
    for s in gl:
        k = lookup(data, s)
        if k is None or k not in con:
            continue
        d = con[k][0]
        hits.append({"gene": k, "log2_phalanx_minus_tibia": round(d, 3)})
        vals.append(d)
    if not vals:
        return None
    return {"n_found": len(vals), "median_log2": round(statistics.median(vals), 3),
            "mean_log2": round(statistics.mean(vals), 3), "genes": hits}


def percentile_of(value, background):
    below = sum(1 for b in background if b < value)
    return round(100.0 * below / len(background), 1)


def load_rat(path, names_path):
    """The rat matrix has a different layout from the mouse one: RefSeq in col 0,
    SYMBOL in col 1, description in col 2, then opaque JL-1891-N_SN library IDs whose
    group membership lives in a separate names workbook (No. -> 'T1wk PZ1' etc.).
    Symbols are UPPERCASE here and Title case in the mouse file."""
    import openpyxl
    wb = openpyxl.load_workbook(names_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    next(it)
    label = {}
    for r in it:
        if r and isinstance(r[0], int) and isinstance(r[1], str):
            label[f"JL-1891-{r[0]}_S{r[0]}"] = r[1].strip()
    wb.close()

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    raw_cols = [str(h) if h is not None else "" for h in hdr[3:]]
    cols = [label.get(c, c) for c in raw_cols]
    data, dropped = {}, 0
    for r in it:
        sym = r[1]
        if not isinstance(sym, str):
            dropped += 1
            continue
        data[sym] = [v if isinstance(v, (int, float)) else None for v in r[3:]]
    wb.close()
    unmapped = [c for c, rc in zip(cols, raw_cols) if c == rc]
    if unmapped:
        raise SystemExit(f"rat: {len(unmapped)} libraries unmapped, e.g. {unmapped[:3]}")
    return cols, data, dropped


def run_species(name, path, names_path=None):
    if name == "rat":
        cols, data, dropped = load_rat(path, names_path)
        g = groups(cols, {"phal_HZ": "Ph1wk HZ", "tib_HZ": "T1wk HZ",
                          "phal_PZ": "Ph1wk PZ", "tib_PZ": "T1wk PZ"})
        for k, v in g.items():
            if not v:
                raise SystemExit(f"rat: group {k} empty; columns were {sorted(set(cols))[:8]}")
        res_pre = (cols, data, dropped, g)
    else:
        cols, data, dropped = load(path)
        g = groups(cols, {"phal_HZ": "1wPh_HZ", "tib_HZ": "1wT_HZ",
                          "phal_PZ": "1wP_PZ", "tib_PZ": "1wT_PZ"})
        res_pre = (cols, data, dropped, g)
    cols, data, dropped, g = res_pre
    res = {"species": name, "genes_loaded": len(data), "excel_corrupted_dropped": dropped,
           "group_n": {k: len(v) for k, v in g.items()}, "zones": {}}

    for zone, (a, b) in {"HZ": ("phal_HZ", "tib_HZ"), "PZ": ("phal_PZ", "tib_PZ")}.items():
        if not g[a] or not g[b]:
            continue
        con = contrast(data, g[a], g[b])       # positive = UP in the SLOW (phalanx) site
        background = [v[0] for v in con.values()]
        bg_median = statistics.median(background)
        zone_out = {"n_genes_in_contrast": len(background),
                    "background_median_log2": round(bg_median, 3),
                    "modules": {}}
        for mod, gl in MODULES.items():
            s = score(con, data, gl)
            if s is None:
                continue
            s["percentile_of_module_median"] = percentile_of(s["median_log2"], background)
            s["vs_background_median"] = round(s["median_log2"] - bg_median, 3)
            zone_out["modules"][mod] = s
        res["zones"][zone] = zone_out
    return res


def main():
    out = {"note": __doc__.strip(),
           "sign_convention": "positive log2 = HIGHER IN THE PHALANX, i.e. in the SLOW, "
                              "matrix-majority plate",
           "species": {}}
    for name, fn, nm in (("mouse", "GSE114919_Mouse_normalizedcounts.xlsx", None),
                         ("rat", "GSE114919_Rat_normalizedcounts.xlsx",
                          "GSE114919_Rat_RNA-Seq_names.xlsx")):
        p = os.path.join(BASE, fn)
        if os.path.exists(p):
            out["species"][name] = run_species(
                name, p, os.path.join(BASE, nm) if nm else None)

    os.makedirs(OUT, exist_ok=True)
    with open(os.path.join(OUT, "vm_module_site_axis.json"), "w") as fh:
        json.dump(out, fh, indent=1)

    for sp, r in out["species"].items():
        print(f"\n########## {sp.upper()}  ({r['genes_loaded']} genes) ##########")
        for zone, z in r["zones"].items():
            print(f"\n--- {zone}: phalanx (slow) minus tibia (fast), "
                  f"{z['n_genes_in_contrast']} genes, background median "
                  f"{z['background_median_log2']:+.3f} ---")
            rows = sorted(z["modules"].items(), key=lambda kv: -kv[1]["median_log2"])
            for mod, s in rows:
                print(f"  {mod:26s} n={s['n_found']:2d}  median {s['median_log2']:+7.3f}  "
                      f"vs bg {s['vs_background_median']:+7.3f}  pct {s['percentile_of_module_median']:5.1f}")


if __name__ == "__main__":
    main()
