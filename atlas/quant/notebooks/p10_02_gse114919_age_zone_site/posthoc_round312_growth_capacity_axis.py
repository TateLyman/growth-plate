#!/usr/bin/env python3
"""
ROUND 312 POST-HOC. The decision this notebook exists to inform is "what should be ADDED to the
stack", and it supplies a discriminator the atlas has never had.

THE DISCRIMINATOR. For any candidate target, two independent contrasts say whether it behaves
like a BRAKE THAT ACCUMULATES or like CAPACITY THAT IS BEING LOST:

  * A target that RISES with age AND is HIGHER in the low-output (finger) plate is behaving like
    a restraint that builds up as growth capacity is spent. Inhibiting it is coherent.
  * A target that FALLS with age AND is LOWER in the low-output plate is capacity being lost.
    Inhibiting it further runs the wrong way; the coherent move there would be to supply it.
  * A target flat on both axes is not tracking growth capacity at all, which is not a
    disqualification - it may act on a term that does not vary this way - but it removes an
    argument people reach for.

WHY BOTH AXES ARE NEEDED. Age alone is the trap CLAUDE.md names: a declining axis looks like a
brake to release and may just be a band. Site alone is confounded by everything else that
separates a phalanx from a tibia. Agreement between two contrasts that share no samples on one
side and no time point on the other is what makes the read worth anything, and DISAGREEMENT
between them is informative too - it says the gene tracks age or size but not growth capacity.

RAT REPLICATION. BOTH contrasts are repeated in rat - T1wk vs T4wk tibia AND Ph1wk phalanx -
in both zones. A mouse effect that does not replicate in rat is downgraded on the spot, and the
genome-wide intersection below REQUIRES rat concordance on both axes wherever rat data exist.

NOTHING HERE IS A PROMOTION. Expression tracking growth capacity is an association in normal
tissue with no perturbation and no length endpoint. It ranks where to look and it can EMBARRASS
a candidate whose direction runs backwards, which is the more valuable of its two uses.
"""
import os
import sys
import json

sys.path.insert(0, os.path.dirname(__file__))
from analysis import load, groups, contrast, MOUSE, RAT  # noqa: E402

OUT = os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "round312")

# Candidate additions and the stack, with the atlas's stated therapeutic direction.
TARGETS = {
    # --- live or recently live candidate ADDITIONS ---
    "Loxl2": "INHIBIT (pLoF +1.40 cm; PXS-5505)", "Plod2": "INHIBIT (pLoF +1.34 cm)",
    "Plod1": "INHIBIT (pLoF +1.03 cm)", "Lox": "INHIBIT (pan-LOX family)",
    "Loxl4": "INHIBIT (pan-LOX family)",
    "Tnks": "INHIBIT (lower canonical Wnt)", "Tnks2": "INHIBIT (lower canonical Wnt)",
    "Axin2": "readout of canonical Wnt", "Ctnnb1": "canonical Wnt effector",
    "Tet1": "INHIBIT (pLoF +8.32 cm)", "Amd1": "INHIBIT? sign unresolved (pLoF +7.18 cm)",
    "Odc1": "polyamine pathway (DFMO target)", "Smox": "polyamine catabolism",
    "Hhip": "SUPPLY/potentiate (pLoF +9.92 cm) - secreted Hh antagonist",
    "Nrk": "INHIBIT (pLoF +3.79 cm)", "Spin4": "reduce (pool lever)",
    "Cxxc5": "present but wrong half (raises Wnt if inhibited)",
    "Npr3": "block clearance", "Fbn1": "perichondrial TGF-beta restraint",
    "Scube3": "SUPPLY (secreted BMP modulator, loss shortens)",
    "Chd8": "INHIBIT (pLoF +10.22 cm)", "Brd2": "INHIBIT (screen delayer)",
    "Kdm1a": "INHIBIT (H3K4 arm)", "Prkar1a": "cAMP band",
    # --- the stack itself, as positive controls for the contrasts ---
    "Fgfr3": "erdafitinib target", "Nppc": "vosoritide ligand", "Npr2": "vosoritide receptor",
    "Cyp19a1": "anastrozole target", "Esr1": "oestrogen/period", "Esr2": "oestrogen",
    "Ghr": "GH arm", "Igf1": "GH arm", "Socs2": "GH arm brake", "Igf1r": "GH arm",
    # --- pathway landmarks ---
    "Ihh": "Hh ligand", "Ptch1": "Hh receptor", "Sufu": "Hh intracellular brake",
    "Gli1": "Hh readout", "Pthlh": "PTHrP", "Pth1r": "PTHrP receptor",
    "Sox9": "chondrocyte TF", "Runx2": "hypertrophy driver", "Acan": "aggrecan",
    "Col2a1": "matrix", "Col10a1": "hypertrophy marker", "Mmp13": "remodelling",
    "Ezh2": "PRC2", "Dot1l": "H3K79", "Aldh2": "screen accelerator",
    "Fkbp1a": "screen accelerator / rapalog binder",
}


def fmt(c, g):
    v = c.get(g)
    return "  n/a  " if v is None else "%+6.2f" % v[0]


def classify(age, site):
    """age/site are log2 differences: age = 4wk-1wk tibia, site = finger-tibia at 1wk."""
    if age is None or site is None:
        return "incomplete"
    T = 0.5
    if age > T and site > T:
        return "BRAKE-LIKE (rises with age, higher in low-output plate)"
    if age < -T and site < -T:
        return "CAPACITY-LIKE (falls with age, lower in low-output plate)"
    if abs(age) <= T and abs(site) <= T:
        return "flat on both"
    return "discordant"


def main():
    cols, data, dropped = load(MOUSE)
    g = groups(cols, {"1wPh_HZ": "1wPh_HZ", "1wT_HZ": "1wT_HZ", "4wT_HZ": "4wT_HZ",
                      "1wPh_PZ": "1wP_PZ", "1wT_PZ": "1wT_PZ", "4wT_PZ": "4wT_PZ"})
    age_pz = contrast(data, g["4wT_PZ"], g["1wT_PZ"])
    age_hz = contrast(data, g["4wT_HZ"], g["1wT_HZ"])
    site_pz = contrast(data, g["1wPh_PZ"], g["1wT_PZ"])
    site_hz = contrast(data, g["1wPh_HZ"], g["1wT_HZ"])

    # RAT. The counts matrix is keyed by library ID (JL-1891-N_SN) and the mapping to
    # condition lives in the separate names workbook, so it is read rather than guessed.
    # The rat replicates BOTH contrasts - T1wk/T4wk and Ph1wk - not only age.
    import openpyxl
    nb = openpyxl.load_workbook(os.path.join(os.path.dirname(MOUSE),
                                             "GSE114919_Rat_RNA-Seq_names.xlsx"), read_only=True)
    ns = nb[nb.sheetnames[0]]
    idx2name = {r[0]: r[1].strip() for r in ns.iter_rows(values_only=True)
                if isinstance(r[0], int) and isinstance(r[1], str)}
    nb.close()
    rcols, rdata, rdrop = load(RAT)

    def rat_group(prefix):
        out = []
        for i, c in enumerate(rcols):
            try:
                n = int(c.split("-")[2].split("_")[0])
            except (IndexError, ValueError):
                continue
            if idx2name.get(n, "").startswith(prefix):
                out.append(i)
        return out

    rg = {"r1_PZ": rat_group("T1wk PZ"), "r4_PZ": rat_group("T4wk PZ"),
          "r1_HZ": rat_group("T1wk HZ"), "r4_HZ": rat_group("T4wk HZ"),
          "rPh_PZ": rat_group("Ph1wk PZ"), "rPh_HZ": rat_group("Ph1wk HZ")}
    rage_pz = contrast(rdata, rg["r4_PZ"], rg["r1_PZ"])
    rage_hz = contrast(rdata, rg["r4_HZ"], rg["r1_HZ"])
    rsite_pz = contrast(rdata, rg["rPh_PZ"], rg["r1_PZ"])
    rsite_hz = contrast(rdata, rg["rPh_HZ"], rg["r1_HZ"])
    print("RAT %d genes, %d libraries; group sizes %s"
          % (len(rdata), len(rcols), {k: len(v) for k, v in rg.items()}))
    print()

    print("=" * 108)
    print("TARGETS - log2 differences.  AGE = 4wk minus 1wk tibia.  SITE = finger minus tibia at 1wk.")
    print("  positive AGE = rises as the window closes | positive SITE = higher in the LOW-OUTPUT plate")
    print("=" * 108)
    print("  %-9s %7s %7s %7s %7s | %7s %7s | %s" %
          ("gene", "agePZ", "ageHZ", "sitePZ", "siteHZ", "ratPZ", "ratHZ", "class / atlas direction"))
    rows = []
    for t in sorted(TARGETS):
        a_pz = age_pz.get(t, (None,))[0]
        a_hz = age_hz.get(t, (None,))[0]
        s_pz = site_pz.get(t, (None,))[0]
        s_hz = site_hz.get(t, (None,))[0]
        cls = classify(a_pz if a_pz is not None else a_hz,
                       s_pz if s_pz is not None else s_hz)
        rows.append({"gene": t, "age_pz": a_pz, "age_hz": a_hz, "site_pz": s_pz, "site_hz": s_hz,
                     "rat_pz": rage_pz.get(t, (None,))[0], "rat_hz": rage_hz.get(t, (None,))[0],
                     "rat_site_pz": rsite_pz.get(t, (None,))[0], "rat_site_hz": rsite_hz.get(t, (None,))[0],
                     "class": cls, "atlas_direction": TARGETS[t]})
        print("  %-9s %s %s %s %s | %s %s | %s -- %s" %
              (t, fmt(age_pz, t), fmt(age_hz, t), fmt(site_pz, t), fmt(site_hz, t),
               fmt(rage_pz, t), fmt(rage_hz, t), cls.split(" (")[0], TARGETS[t]))

    # genome-wide intersection
    print()
    print("=" * 108)
    print("GENOME-WIDE INTERSECTION - proliferative zone")
    print("=" * 108)
    brake, cap = [], []
    for gn in age_pz:
        if gn not in site_pz:
            continue
        a, s = age_pz[gn][0], site_pz[gn][0]
        r = rage_pz.get(gn, (None,))[0]
        rs = rsite_pz.get(gn, (None,))[0]
        if a > 1.0 and s > 1.0 and (r is None or r > 0) and (rs is None or rs > 0):
            brake.append((gn, a, s, r))
        if a < -1.0 and s < -1.0 and (r is None or r < 0) and (rs is None or rs < 0):
            cap.append((gn, a, s, r))
    brake.sort(key=lambda x: -(x[1] + x[2]))
    cap.sort(key=lambda x: (x[1] + x[2]))
    print("  BRAKE-LIKE (up with age AND up in low-output plate, rat-concordant where available): %d"
          % len(brake))
    for gn, a, s, r in brake[:35]:
        print("    %-14s age %+6.2f  site %+6.2f  rat %s" % (gn, a, s, "%+.2f" % r if r is not None else "n/a"))
    print("  CAPACITY-LIKE (down with age AND down in low-output plate): %d" % len(cap))
    for gn, a, s, r in cap[:35]:
        print("    %-14s age %+6.2f  site %+6.2f  rat %s" % (gn, a, s, "%+.2f" % r if r is not None else "n/a"))

    os.makedirs(OUT, exist_ok=True)
    json.dump({"targets": rows,
               "brake_like_pz": [{"gene": x[0], "age": x[1], "site": x[2], "rat": x[3]} for x in brake],
               "capacity_like_pz": [{"gene": x[0], "age": x[1], "site": x[2], "rat": x[3]} for x in cap]},
              open(os.path.join(OUT, "gse114919_growth_capacity_axis.json"), "w"), indent=1)
    print()
    print("  wrote gse114919_growth_capacity_axis.json")


if __name__ == "__main__":
    main()
