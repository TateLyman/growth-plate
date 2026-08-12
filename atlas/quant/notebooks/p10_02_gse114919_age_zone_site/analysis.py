#!/usr/bin/env python3
"""
GSE114919 - GROWTH PLATE BY AGE x ZONE x SITE, mouse and rat. Operator-supplied 2026-08-13.

WHY THIS IS THE RIGHT DATASET FOR THE BINDING CONSTRAINT. At bone age 16 the scarce resource
is the PERIOD, not the rate. Every other transcriptome in this atlas is a snapshot of one plate
at one age; this one carries TWO orthogonal contrasts of growth capacity, zone-resolved, in two
species:

  AGE   1 week vs 4 week TIBIA, in proliferative and hypertrophic zones separately.
        A 1-week mouse plate is fast and has its whole period ahead of it; a 4-week plate is
        well into decline. This is the closest thing the atlas has to "what changes as the
        window closes."
  SITE  1 week TIBIA vs 1 week FINGER (phalanx), same animals, same age, same zones.
        A phalangeal plate and a tibial plate are both open and both growing, but the tibia
        will contribute an order of magnitude more length. This is "what distinguishes a
        high-output plate from a low-output one" with age held fixed.

The INTERSECTION is the point. A gene that falls with age AND is lower in the low-output site
is tracking growth capacity on two independent axes, which is a far stronger signal than either
contrast alone - and neither contrast can be explained by the other.

PREREGISTERED WARNING, WRITTEN BEFORE LOOKING, because CLAUDE.md names this exact failure.
"Any axis found DECLINING with maturation looks like a brake to release and may just be
regulated within a band." FIVE axes in this atlas already shorten bone in BOTH directions.
A gene that declines with age is a CANDIDATE and not a lever, and this notebook cannot promote
anything on its own - it can only rank where to look. The site contrast is also confounded by
everything else that differs between a phalanx and a tibia (HOX code, absolute size, mechanical
environment, vascular supply), so it is used only as a second axis for intersection, never alone.

DATA NOTE. The finger proliferative-zone group is named 1wP_PZ in the header (not 1wPh_PZ) and
has FOUR replicates, not five - sample 1wP_PZ3 is absent from the deposited matrix. Every other
group has five. The supplied matrix has 20,181 rows of which 26 gene symbols were destroyed by
Excel date auto-conversion (the MARCH* and SEPT* families appear as serial numbers such as
41333). Those rows are DROPPED rather than guessed at. Values are already normalised and
log-scaled by the depositors.

Usage: python3 analysis.py
"""
import os
import math
import openpyxl

BASE = os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "supplied_2026_08_13")
MOUSE = os.path.join(BASE, "GSE114919_Mouse_normalizedcounts.xlsx")
RAT = os.path.join(BASE, "GSE114919_Rat_normalizedcounts.xlsx")


def load(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    cols = [str(h) if h is not None else "" for h in hdr]
    data, dropped = {}, 0
    for r in it:
        g = r[0]
        if not isinstance(g, str):
            dropped += 1          # Excel date-corrupted symbol; not guessed at
            continue
        vals = [v if isinstance(v, (int, float)) else None for v in r[1:]]
        data[g] = vals
    wb.close()
    return cols[1:], data, dropped


def groups(cols, spec):
    """spec: {label: substring} -> {label: [column indices]}"""
    out = {}
    for lab, sub in spec.items():
        out[lab] = [i for i, c in enumerate(cols) if c.startswith(sub)]
    return out


def mean(vals, idx):
    xs = [vals[i] for i in idx if i < len(vals) and vals[i] is not None]
    return sum(xs) / len(xs) if xs else None


def contrast(data, gA, gB, min_expr=1.0):
    """log2 difference of already-log-scaled group means, A minus B."""
    out = {}
    for g, v in data.items():
        a, b = mean(v, gA), mean(v, gB)
        if a is None or b is None:
            continue
        if max(a, b) < min_expr:
            continue          # both groups effectively silent; a ratio there is noise
        out[g] = (a - b, a, b)
    return out


def main():
    cols, data, dropped = load(MOUSE)
    print("MOUSE %d genes (%d Excel-corrupted symbols dropped), %d libraries"
          % (len(data), dropped, len(cols)))
    print("  columns:", ", ".join(sorted(set(c.rsplit("_", 1)[0] if "_" in c else c for c in cols))))

    g = groups(cols, {"1wPh_HZ": "1wPh_HZ", "1wT_HZ": "1wT_HZ", "4wT_HZ": "4wT_HZ",
                      "1wPh_PZ": "1wP_PZ", "1wT_PZ": "1wT_PZ", "4wT_PZ": "4wT_PZ"})
    for k, v in g.items():
        print("  group %-8s n=%d" % (k, len(v)))

    age_pz = contrast(data, g["4wT_PZ"], g["1wT_PZ"])   # positive = UP with age
    age_hz = contrast(data, g["4wT_HZ"], g["1wT_HZ"])
    site_pz = contrast(data, g["1wPh_PZ"], g["1wT_PZ"])  # positive = UP in the LOW-output site
    site_hz = contrast(data, g["1wPh_HZ"], g["1wT_HZ"])
    return dict(cols=cols, data=data, groups=g,
                age_pz=age_pz, age_hz=age_hz, site_pz=site_pz, site_hz=site_hz)


if __name__ == "__main__":
    main()
